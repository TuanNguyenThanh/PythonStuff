import openpyxl
from openpyxl import Workbook
import xlwings as xw
import init as init
import os
import glob

# Ham cham bai thi Excel
# Function name: markTest()
# Input: testFile 
# Output: Array: [testFile,markSum,markDetail[]]
def markTest(testFile):    
    tempFile = init.tempFile
    answerFile = init.answerFile
    questions = init.questions
    print('1. load answer workbook & sheet')
    answerWb = openpyxl.load_workbook(answerFile,data_only=True)
    answerSheet = answerWb.active
    print('2. load temp workbook & sheet')
    tempWb = openpyxl.load_workbook(tempFile)
    tempSheet = tempWb.active
    print('3. load workbook & sheet testFile')
    wb = openpyxl.load_workbook(testFile)
    sheet = wb.active
    print('4. copy answers cell to tempFile')
    for i in range(len(questions)):
        print('Question ', i)
        # print('-------')
        for cell in sheet[questions[i][0]]:
            for j in range(len(cell)):
                # print(j)
                s = cell[j].value
                # print(s)
                # print(s)
                if isinstance(s, str) and s[0] == '=':
                    tempSheet.cell(row = cell[j].row,column = cell[j].column).value = cell[j].value
                else:
                    tempSheet.cell(row = cell[j].row,column = cell[j].column).value = '?'
    
    print('5. Save temp workbook, calculate the formulas and resave tempFile')
    tempWb.save(tempFile)
    tempWbook = xw.Book(tempFile)
    tempWsheet = tempWbook.sheets.active
    app = xw.apps.active 
    print('6. Compare tempFile value with answer & Mark test')
    sum = 0
    markDetail = [testFile]
    for i in range(len(questions)):
        point = questions[i][1] / len(answerSheet[questions[i][0]])
        if len(answerSheet[questions[i][0]]) == 1:
            point = point / len(answerSheet[questions[i][0]][0])
        # point_div = 0
        # for cell in answerSheet[questions[i][0]]:
        #     for j in range(len(cell)):
        #         point_div = point_div + 1
        # print(point_div)
        count = 0
        for cell in answerSheet[questions[i][0]]:
            for j in range(len(cell)):
                answerAdd = cell[j].column_letter + str(cell[j].row)
                # address = cell[j].column_letter + str(cell[j].row)
                # print(addres s)
                # print(tempWsheet.range(address).value)
                if tempWsheet.range(answerAdd).value == answerSheet[answerAdd].value:
                    count = count + point
        count = round(count,1)
        markDetail.append(count)
        sum = sum + count
        print(count)
    print('Mark: :', round(sum,1))
    # app.quit()
    tempWbook.close()
    markDetail.insert(1,sum)
    return markDetail
    # return resual

def markTestNew(testFile):    
    tempFiles = init.tempFiles
    answerFiles = init.answerFiles
    questions = init.questions
    markDetail = [1000 for i in range(len(questions))] #Save detail result
    
    print('1. load workbook & sheet testFile')
    wb = openpyxl.load_workbook(testFile)
    sheet = wb.active
    
    print('2. load answer workbooks & sheets')
    for a in range(len(answerFiles)):
        print('----------------',' test ',a,'-----------------')
        answerWb = openpyxl.load_workbook(answerFiles[a],data_only=True)
        answerSheet = answerWb.active
        
        print('3. load temp workbook & sheet')
        tempWb = openpyxl.load_workbook(tempFiles[a])
        tempSheet = tempWb.active
        
        print('4. copy answers cell to tempFile')
        for i in range(len(questions)):
            print('Question ', i)
            # print('-------')
            for cell in sheet[questions[i][0]]:
                for j in range(len(cell)):
                    s = cell[j].value
                    if isinstance(s, str) and s[0] == '=':
                        tempSheet.cell(row = cell[j].row,column = cell[j].column).value = cell[j].value
                    else:
                        tempSheet.cell(row = cell[j].row,column = cell[j].column).value = '?'
        
        print('5. Save temp workbook, calculate the formulas and resave tempFile')
        tempWb.save(tempFiles[a])
        tempWbook = xw.Book(tempFiles[a])
        tempWsheet = tempWbook.sheets.active
        # app = xw.apps.active 
        
        print('6. Compare tempFile value with answer & Mark test')
        for i in range(len(questions)):
            point = questions[i][1] / len(answerSheet[questions[i][0]])
            if len(answerSheet[questions[i][0]]) == 1:
                point = point / len(answerSheet[questions[i][0]][0])
            # point_div = 0
            # for cell in answerSheet[questions[i][0]]:
            #     for j in range(len(cell)):
            #         point_div = point_div + 1
            # print(point_div)
            count = 0
            for cell in answerSheet[questions[i][0]]:
                for j in range(len(cell)):
                    answerAdd = cell[j].column_letter + str(cell[j].row)
                    if tempWsheet.range(answerAdd).value == answerSheet[answerAdd].value:
                        count = count + point
            count = round(count,1)
            if count < markDetail[i]:
                markDetail[i] = count
            print(count)
        tempWbook.close()
    sum = 0
    for i in range(len(markDetail)):
        markDetail[i] = markDetail[i] % 1000
        sum = sum + markDetail[i]
    markDetail.insert(0,testFile)
    markDetail.insert(1,round(sum,1))
    print(markDetail)
    return markDetail
    # return resual

def markTestNew2(testFile):    
    tempFiles = init.tempFiles
    answerFiles = init.answerFiles
    questions = init.questions
    markDetail = [1000 for i in range(len(questions))] #Save detail result
    # print('1. load workbook & sheet testFile')
    wb = openpyxl.load_workbook(testFile)
    sheet = wb.active

    # print('2. load answer workbooks & sheets')
    for a in range(len(answerFiles)):
        print('----------------',' test ',a + 1,'-----------------')
        answerWb = openpyxl.load_workbook(answerFiles[a],data_only=True)
        answerSheet = answerWb.active
        
        # print('3. load temp workbook & sheet')
        # tempWb = openpyxl.load_workbook(tempFiles[a])
        # tempSheet = tempWb.active
        # print('4. copy test cell to tempFile')
        for i in range(len(questions)):
            tempWbook = xw.Book(tempFiles[a])
            tempWsheet = tempWbook.sheets.active 
            app = xw.apps.active 
            for cell in sheet[questions[i][0]]:
                for j in range(len(cell)):
                    s = cell[j].value
                    answerAdd = cell[j].column_letter + str(cell[j].row)
                    if isinstance(s, str) and s[0] == '=':
                        tempWsheet.range(answerAdd).value = s
                    else:
                        tempWsheet.range(answerAdd).value = '?'
                    # print(tempSheet.cell(row = cell[j].row,column = cell[j].column).value)
            # print('5. Save temp workbook, calculate the formulas and resave tempFile')
            tempWbook.save()
            #calculate the formulas and resave tempFile
            # tempWbook = xw.Book(tempFiles[a])
            # tempWsheet = tempWbook.sheets.active 
            tempWbook.app.calculate()
            # print(tempWsheet.range('H5').value)
            # print('6. Compare tempFile value with answer & Mark test')
            point = questions[i][1] / len(answerSheet[questions[i][0]])
            if len(answerSheet[questions[i][0]]) == 1:
                point = point / len(answerSheet[questions[i][0]][0])
            count = 0
            for cell in answerSheet[questions[i][0]]:
                for j in range(len(cell)):
                    answerAdd = cell[j].column_letter + str(cell[j].row)
                    # print('===========',answerAdd,'===========')
                    # print(tempWsheet.range(answerAdd).value)
                    if tempWsheet.range(answerAdd).value == answerSheet[answerAdd].value:
                        count = count + point
                    tempWsheet.range(answerAdd).value = answerSheet[answerAdd].value
                    # print(answerSheet[answerAdd].value)
            count = round(count,1)
            if count < markDetail[i]:
                markDetail[i] = count
            # print(count)
            print('-------Question ', i + 1, ': ',markDetail[i] % 1000)
            tempWbook.save()
            tempWbook.close() #for macos
            # app.kill() #for windows
    sum = 0
    for i in range(len(markDetail)):
        markDetail[i] = markDetail[i] % 1000
        sum = sum + markDetail[i]
    markDetail.insert(0,testFile)
    markDetail.insert(1,round(sum,1))
    print(markDetail)
    return markDetail
    # return resual


# Ham cham nhieu bai thi Excel
# Function name: markMultiTest()
# Input: none 
# Output: resultFile.xlsx
def markMultiTest():
    # init program 
    nameFolder = init.nameFolder
    resualFile = init.resualFile
    records = init.records
    #Create file result.xlsx
    wb = Workbook()
    ws = wb.active
    for i in range(len(records)):
        ws.cell(row = 1, column = i + 1, value = records[i])
    # read files from Folder
    listFolder = glob.glob(nameFolder + "/*")
    count = 0
    for folder in listFolder:
        print('____________________')
        print('Bai thi ', count)
        folderName = folder.replace(nameFolder + '/','')
        # roomName = folderName.split('-')[0]
        # no = int(folderName.split('-')[1])
        # name = folderName.split('-')[2].split('_')[0]
        count = count + 1
        # print(roomName,' ',no,' ',name)
        name = folderName.split('_')[0]
        sbd = folderName.split('_')[1]
        f = glob.glob(folder+'/*.xlsx')
        if not f:
            print ('Loi ten file')
            f = glob.glob(folder+'/*.XLSX')
        if f:
            try:
                result = markTestNew2(f.pop())
            except:
                print('Loi file')
                result = []
        else:
            result = []
        # ws.cell(row = count + 1, column = 1, value = roomName)
        ws.cell(row = count + 1, column = 1, value = sbd)
        ws.cell(row = count + 1, column = 2, value = name)
        if len(result) != 0 :
            for i in range(len(result)):
                ws.cell(row = count + 1, column = i + 3, value = result[i])
            print('OK')
        else:
            print('no file')
            ws.cell(row = count + 1, column = 3, value = 'no file')
            ws.cell(row = count + 1, column = 4, value = 0) 
        # wb = openpyxl.load_workbook(f.pop())
        # sheet = wb.active;
    wb.save(resualFile)
# markMultiTest()
# markTest('kt.xlsx')
markTestNew2('ldtp.xlsx')
markTestNew2('ard.xlsx')