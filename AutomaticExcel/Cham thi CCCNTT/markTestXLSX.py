import openpyxl
from openpyxl import Workbook
import xlwings as xw
import init as init
import sys
import os
import glob

# Ham cham bai thi Excel
# Function name: markTest()
# Input: testFile 
# Output: Array: [testFile,markSum,markDetail[]]
def markTest(testFile):    
    tempFiles = init.tempFiles
    answerFiles = init.answerFiles
    questions = init.questions
    markDetail = [1000 for i in range(len(questions))] #Save detail result
    # print('1. load workbook & sheet testFile')
    wb = openpyxl.load_workbook(testFile)
    sheet = wb.active

    # print('2. load answer workbooks & sheets')
    for a in range(len(answerFiles)):
        print('----------------',' test ',a + 1,'-----------------')
        answerWb = openpyxl.load_workbook(answerFiles[a],data_only=True)
        answerSheet = answerWb.active
        # print('3. copy test cell to tempFile, calculate the formulas and resave')
        for i in range(len(questions)):
            # print('3.1 load temp workbook & sheet')
            tempWbook = xw.Book(tempFiles[a])
            tempWsheet = tempWbook.sheets.active 
            app = xw.apps.active 
            for cell in sheet[questions[i][0]]:
                for j in range(len(cell)):
                    s = cell[j].value
                    answerAdd = cell[j].column_letter + str(cell[j].row)
                    if isinstance(s, str) and s[0] == '=':
                        tempWsheet.range(answerAdd).value = s
                    else:
                        tempWsheet.range(answerAdd).value = '?'
            # print('3.2 Save temp workbook, calculate the formulas and resave tempFile')
            tempWbook.save()
            #calculate the formulas
            tempWbook.app.calculate()
            # print('3.3 Compare tempFile value with answer & Mark test')
            point = questions[i][1] / len(answerSheet[questions[i][0]])
            if len(answerSheet[questions[i][0]]) == 1:
                point = point / len(answerSheet[questions[i][0]][0])
            count = 0
            for cell in answerSheet[questions[i][0]]:
                for j in range(len(cell)):
                    answerAdd = cell[j].column_letter + str(cell[j].row)
                    if tempWsheet.range(answerAdd).value == answerSheet[answerAdd].value:
                        count = count + point
                    tempWsheet.range(answerAdd).value = answerSheet[answerAdd].value
            count = round(count,1)
            if count < markDetail[i]:
                markDetail[i] = count
            print('-------Question ', i + 1, ': ',markDetail[i] % 1000)
            tempWbook.save()
            tempWbook.close() #for macos
            # app.kill() #for windows
    sum = 0
    for i in range(len(markDetail)):
        markDetail[i] = markDetail[i] % 1000
        sum = sum + markDetail[i]
    markDetail.insert(0,testFile)
    markDetail.insert(1,round(sum,1))
    print(markDetail)
    return markDetail

def markMultiTest():
    # init program 
    nameFolder = init.nameFolder
    resualFile = init.resualFile
    records = init.records
    #Create file result.xlsx
    wb = Workbook()
    ws = wb.active
    for i in range(len(records)):
        ws.cell(row = 1, column = i + 1, value = records[i])
    # read files from Folder
    listFolder = glob.glob(nameFolder + "/*")
    count = 0
    for folder in listFolder:
        print('____________________')
        print('Bai thi ', count)
        folderName = folder.replace(nameFolder + '/','')
        # roomName = folderName.split('-')[0]
        # no = int(folderName.split('-')[1])
        # name = folderName.split('-')[2].split('_')[0]
        count = count + 1
        # print(roomName,' ',no,' ',name)
        name = folderName.split('_')[0]
        sbd = folderName.split('_')[1]
        f = glob.glob(folder+'/*.xlsx')
        if not f:
            print ('Loi ten file')
            f = glob.glob(folder+'/*.XLSX')
        if f:
            try:
                result = markTestNew2(f.pop())
            except:
                print('Loi file')
                result = []
        else:
            result = []
        # ws.cell(row = count + 1, column = 1, value = roomName)
        ws.cell(row = count + 1, column = 1, value = sbd)
        ws.cell(row = count + 1, column = 2, value = name)
        if len(result) != 0 :
            for i in range(len(result)):
                ws.cell(row = count + 1, column = i + 3, value = result[i])
            print('OK')
        else:
            print('no file')
            ws.cell(row = count + 1, column = 3, value = 'no file')
            ws.cell(row = count + 1, column = 4, value = 0) 
        # wb = openpyxl.load_workbook(f.pop())
        # sheet = wb.active;
    wb.save(resualFile)

if __name__ == '__main__':
    globals()[sys.argv[1]](sys.argv[2])