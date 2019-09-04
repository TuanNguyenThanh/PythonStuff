import openpyxl
import glob
import os

def replace(fileName,oldValue,newValue,cellsName):
    excelFile = openpyxl.load_workbook(fileName)
    sheet = excelFile.active
    for cell in sheet[cellsName]:
        for j in range(len(cell)):
                # print(j)
                s = cell[j].value
                cell[j].value = s.replace(oldValue,newValue,1)
                # print(cell[j].value)
    excelFile.save(fileName)
    # currentRow = 1
    # for eachRow in sheet1.iter_rows():
    #     sheet1.cell(row=currentRow, column=2).value = "0"
    #     currentRow += 1
    # excelFile.save(fileName)
def replaceMultiFile(folderName):
    listFolder = glob.glob(folderName + "/*")
    count = 0
    for folder in listFolder:
        # print('____________________')
        f = glob.glob(folder+'/*.xlsx')
        if not f: 
            print ('Loi ten file')
            f = glob.glob(folder+'/*.XLSX')
            if not f:
                print ('Tiep tuc Loi ten file')
        try:
            fileName = f.pop()
            replace(fileName,'HONGKONG','HONGKHONG','C3:C11')
            print('File ',count + 1,' ',fileName,' - success')
            count = count + 1
        except:
            # print('File ',count + 1,' ',fileName,' - error')
            count = count + 1
            print(folder)
            print('Loi file ')
# replace('test.xlsx','abc','C3:C11')

def renameMultiFile(folderName, newFileName):
    listFolder = glob.glob(folderName + "/*")
    count = 0
    success = 0
    error = 0
    for folder in listFolder:
        count = count + 1
        # success 
        # print('____________________')
        f = glob.glob(folder+'/*.*')
        print(folder)
        if not f: 
            print('Loi ten file')
        try:
            fileName = f.pop()
            os.rename(fileName,folder+'/'+newFileName)
            success = success + 1
        except OSError as e:
            error = error + 1
            print('Loi mo file: ',e)
    print('Doi ten ',count,' file, thanh cong: ', success,', that bai: ',error)
# replaceMultiFile('Ca1')
renameMultiFile('Ca2','de2.xlsx')