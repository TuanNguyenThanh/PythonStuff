import openpyxl
from pycel import ExcelCompiler
import os
import glob
##############config############
nameFolder = 'test'
tempFile = 'De01.xlsx'
answerFile = 'De01answer.xlsx'
resualFile = 'resual.xlsx'
question = [['H5:H14',1], ['I5:I14',1], ['D5:D14',1], ['M5:M9',1], ['L14:M14',1]]

# ##############################
listFolder = glob.glob(nameFolder + "/*")
answerWb = openpyxl.load_workbook(answerFile,data_only=True)
answerSheet = answerWb.active
tempWb = openpyxl.load_workbook(tempFile)
tempSheet = tempWb.active

resualSheet = openpyxl.load_workbook(resualFile).active;
resualSheet['A1'] =  'No'
resualSheet['A2'] = 'Name'
resualSheet['A3'] = 'Prename'

for folder in listFolder:
# folder = listFolder[0]
    name = folder.replace(nameFolder + '/','').split('_')[0]
    print(name)
    f = glob.glob(folder+'/*.xlsx')
        # print(f)
    wb = openpyxl.load_workbook(f.pop())
    sheet = wb.active;

    ## copy value from bl to tempfile
    for i in range(len(question)):
        if i == 5:
            for cell in sheet[question[i][0]]:
                s = cell[0].value
                if isinstance(s, str) and s[0] == '=':
                    tempSheet.cell(row = cell[0].row,column = cell[0].column).value = cell[0].value
                else:
                    tempSheet.cell(row = cell[0].row,column = cell[0].column).value = '?'
    tempWb.save('test.xlsx')

    excel = ExcelCompiler('test.xlsx')
    # excel.from_file()
    # 
    # print(excel.evaluate('Sheet1!H5'))
    excel.to_file('test.xlsx')

    # tempWb = openpyxl.load_workbook('test.xlsx',data_only=True)
    # tempSheet = tempWb.active
    # print(answerSheet['H5'].value)
    sum = 0
    for i in range(5):
        point = question[i][1] / len(answerSheet[question[i][0]]);
        count = 0;
        for cell in answerSheet[question[i][0]]:
            answerAdd = cell[0].column_letter + str(cell[0].row)
            address = answerSheet.title +'!'+ cell[0].column_letter + str(cell[0].row)
            if excel.evaluate(address) == answerSheet[answerAdd].value:
                count = count + point
            # print(excel.evaluate(address))
        sum = sum + round(count,1)
        print(round(count,1))
    print(sum)
